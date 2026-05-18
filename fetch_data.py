import os, requests, json, base64
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, date
import re

# ── Microsoft / SharePoint ─────────────────────────────────────────────
CLIENT_ID       = os.environ["CLIENT_ID"]
TENANT_ID       = os.environ["TENANT_ID"]
CLIENT_SECRET   = os.environ["CLIENT_SECRET"]
REFRESH_TOKEN   = os.environ["REFRESH_TOKEN"]
FILE_URL        = os.environ["SHAREPOINT_FILE_URL"]

# ── Google Drive / CUDYR ───────────────────────────────────────────────
G_CLIENT_ID     = os.environ["GOOGLE_CLIENT_ID"]
G_CLIENT_SECRET = os.environ["GOOGLE_CLIENT_SECRET"]
G_REFRESH_TOKEN = os.environ["GOOGLE_REFRESH_TOKEN"]

# Carpeta CUDYR 2026
CUDYR_2026_FOLDER = "1DdnURFrLf9pSX67J1WzsCDLmOUIQu0D5"

# ── GitHub (auto-renovación tokens) ───────────────────────────────────
GH_PAT  = os.environ.get("GH_PAT", "")
GH_REPO = os.environ.get("GITHUB_REPOSITORY", "")

MESES_VALIDOS = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
                 "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]

MESES_ES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

# ══════════════════════════════════════════════════════════════════════
# MICROSOFT TOKEN
# ══════════════════════════════════════════════════════════════════════
def get_ms_token():
    resp = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "refresh_token": REFRESH_TOKEN,
            "grant_type": "refresh_token",
            "scope": "Files.Read offline_access"
        }
    )
    data = resp.json()
    if "access_token" not in data:
        raise Exception(f"Error MS auth: {data}")
    new_rt = data.get("refresh_token")
    if new_rt and new_rt != REFRESH_TOKEN and GH_PAT and GH_REPO:
        try:
            update_github_secret("REFRESH_TOKEN", new_rt)
            print("✓ MS refresh token renovado")
        except Exception as e:
            print(f"⚠ No se pudo renovar MS token: {e}")
    return data["access_token"]

# ══════════════════════════════════════════════════════════════════════
# GOOGLE TOKEN
# ══════════════════════════════════════════════════════════════════════
def get_google_token():
    resp = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": G_CLIENT_ID,
            "client_secret": G_CLIENT_SECRET,
            "refresh_token": G_REFRESH_TOKEN,
            "grant_type": "refresh_token"
        }
    )
    data = resp.json()
    if "access_token" not in data:
        raise Exception(f"Error Google auth: {data}")
    return data["access_token"]

# ══════════════════════════════════════════════════════════════════════
# GITHUB SECRETS
# ══════════════════════════════════════════════════════════════════════
def update_github_secret(secret_name, secret_value):
    headers = {"Authorization": f"token {GH_PAT}", "Accept": "application/vnd.github.v3+json"}
    r = requests.get(f"https://api.github.com/repos/{GH_REPO}/actions/secrets/public-key", headers=headers)
    r.raise_for_status()
    key_data = r.json()
    from base64 import b64encode
    from nacl import encoding, public
    pk = public.PublicKey(key_data["key"].encode(), encoding.Base64Encoder)
    box = public.SealedBox(pk)
    encrypted = b64encode(box.encrypt(secret_value.encode())).decode()
    r2 = requests.put(
        f"https://api.github.com/repos/{GH_REPO}/actions/secrets/{secret_name}",
        headers=headers,
        json={"encrypted_value": encrypted, "key_id": key_data["key_id"]}
    )
    r2.raise_for_status()

# ══════════════════════════════════════════════════════════════════════
# SHAREPOINT — descargar Excel censo
# ══════════════════════════════════════════════════════════════════════
def get_sharepoint_file(token):
    headers = {"Authorization": f"Bearer {token}"}
    b64 = base64.urlsafe_b64encode(FILE_URL.encode()).decode().rstrip("=")
    url = f"https://graph.microsoft.com/v1.0/shares/u!{b64}/driveItem/content"
    resp = requests.get(url, headers=headers, allow_redirects=True)
    if resp.status_code != 200:
        raise Exception(f"Error descargando Excel: {resp.status_code} - {resp.text[:300]}")
    return resp.content

# ══════════════════════════════════════════════════════════════════════
# GOOGLE DRIVE — listar archivos CUDYR y leer TOTAL MENSUAL
# ══════════════════════════════════════════════════════════════════════
def list_drive_files(token, folder_id):
    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://www.googleapis.com/drive/v3/files?q='{folder_id}'+in+parents+and+mimeType='application/vnd.google-apps.spreadsheet'&fields=files(id,name)&pageSize=50"
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    return resp.json().get("files", [])

def export_sheet_as_csv(token, file_id, sheet_name="TOTAL MENSUAL"):
    """Exporta una hoja específica como CSV usando la Sheets API"""
    headers = {"Authorization": f"Bearer {token}"}
    # Primero obtener el gid de la hoja
    meta_url = f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}?fields=sheets.properties"
    meta = requests.get(meta_url, headers=headers).json()
    
    gid = None
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if props.get("title", "").upper() == sheet_name.upper():
            gid = props.get("sheetId")
            break
    
    if gid is None:
        return None
    
    # Exportar como CSV
    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv&gid={gid}"
    resp = requests.get(export_url, headers=headers)
    if resp.status_code != 200:
        return None
    return resp.text

def parse_cudyr_total_mensual(csv_text):
    """Extrae % cuidados críticos, medios y básicos del TOTAL MENSUAL"""
    if not csv_text:
        return None
    
    lines = csv_text.split('\n')
    
    criticos = medios = basicos = total_pacientes = 0
    
    for i, line in enumerate(lines):
        # Buscar fila con totales: busca patrones de números en filas con encabezados de categoría
        # La fila de resumen tiene: A1,A2,A3 (criticos), B1,B2,B3 (medios), C1,C2,C3 (basicos)
        if 'TOTAL MENSUAL' in line.upper() or ('% CUIDADOS' in line.upper()):
            # Buscar en líneas siguientes
            for j in range(i, min(i+10, len(lines))):
                cols = lines[j].split(',')
                # Buscar línea con % Críticos
                line_up = lines[j].upper()
                if 'CRÍT' in line_up or 'CRITI' in line_up or 'CRÍTICOS' in line_up:
                    nums = [c.strip().rstrip('%') for c in cols if c.strip().rstrip('%').replace('.','').replace(',','').isdigit() or (c.strip().endswith('%') and c.strip()[:-1].replace('.','').isdigit())]
                    if nums:
                        try: criticos = float(nums[0])
                        except: pass
                if 'MEDIOS' in line_up or 'MEDIO' in line_up:
                    nums = [c.strip().rstrip('%') for c in cols if c.strip().rstrip('%').replace('.','').isdigit() or (c.strip().endswith('%') and c.strip()[:-1].replace('.','').isdigit())]
                    if nums:
                        try: medios = float(nums[0])
                        except: pass
                if 'BÁSIC' in line_up or 'BASIC' in line_up or 'BÁSICOS' in line_up:
                    nums = [c.strip().rstrip('%') for c in cols if c.strip().rstrip('%').replace('.','').isdigit() or (c.strip().endswith('%') and c.strip()[:-1].replace('.','').isdigit())]
                    if nums:
                        try: basicos = float(nums[0])
                        except: pass

    # Si no encontramos con el método anterior, usar Sheets API directamente
    return None  # Señal para usar método alternativo

def get_cudyr_via_sheets_api(token, file_id):
    """Usa Sheets API para leer los valores directamente"""
    headers = {"Authorization": f"Bearer {token}"}
    
    # Leer el rango donde está el resumen mensual
    # Basado en la estructura vista: buscar en toda la hoja TOTAL MENSUAL
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}/values/TOTAL%20MENSUAL?majorDimension=ROWS"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        return None
    
    rows = resp.json().get("values", [])
    
    criticos_pct = medios_pct = basicos_pct = None
    total_dias = 0
    criticos_dias = medios_dias = basicos_dias = 0
    
    for i, row in enumerate(rows):
        row_str = ' '.join(str(c) for c in row).upper()
        
        # Buscar fila de totales mensuales con A (críticos), B (medios), C (básicos)
        # Patrón: % CUIDADOS CRÍTICOS / MEDIOS / BÁSICOS
        if '% CUIDADOS CR' in row_str or 'CUIDADOS CR' in row_str:
            for cell in row:
                cell_s = str(cell).strip().replace('%','').replace(',','.')
                try:
                    val = float(cell_s)
                    if 0 < val <= 100:
                        criticos_pct = round(val, 1)
                        break
                except: pass
        
        if '% CUIDADOS ME' in row_str or ('CUIDADOS ME' in row_str and 'CR' not in row_str):
            for cell in row:
                cell_s = str(cell).strip().replace('%','').replace(',','.')
                try:
                    val = float(cell_s)
                    if 0 < val <= 100:
                        medios_pct = round(val, 1)
                        break
                except: pass

        if '% CUIDADOS BÁ' in row_str or '% CUIDADOS BA' in row_str or 'CUIDADOS BÁS' in row_str:
            for cell in row:
                cell_s = str(cell).strip().replace('%','').replace(',','.')
                try:
                    val = float(cell_s)
                    if 0 < val <= 100:
                        basicos_pct = round(val, 1)
                        break
                except: pass

        # Buscar totales de días por categoría (fila TOTAL con números grandes)
        if 'TOTAL' in row_str and len(row) > 5:
            nums = []
            for cell in row:
                try:
                    v = int(str(cell).strip())
                    if v > 0:
                        nums.append(v)
                except: pass
            if len(nums) >= 3:
                # A=críticos, B=medios, C=básicos
                criticos_dias = nums[0] if len(nums) > 0 else 0
                medios_dias   = nums[1] if len(nums) > 1 else 0
                basicos_dias  = nums[2] if len(nums) > 2 else 0
                total_dias    = sum(nums[:3])

    # Si no encontramos porcentajes calculados, los calculamos nosotros
    if criticos_pct is None and total_dias > 0:
        criticos_pct = round(criticos_dias / total_dias * 100, 1)
        medios_pct   = round(medios_dias   / total_dias * 100, 1)
        basicos_pct  = round(basicos_dias  / total_dias * 100, 1)

    if criticos_pct is None:
        return None

    return {
        "criticos_pct": criticos_pct,
        "medios_pct":   medios_pct   or 0,
        "basicos_pct":  basicos_pct  or 0,
        "total_dias":   total_dias,
        "criticos_dias": criticos_dias,
        "medios_dias":   medios_dias,
        "basicos_dias":  basicos_dias,
    }

def extract_month_from_name(name):
    """Extrae el mes del nombre del archivo CUDYR"""
    name_lower = name.lower()
    for mes, num in MESES_ES.items():
        if mes in name_lower:
            return mes.upper(), num
    return None, None

def fetch_cudyr_data(g_token):
    """Lee todos los archivos CUDYR 2026 y extrae los datos de TOTAL MENSUAL"""
    print("  Listando archivos CUDYR 2026...")
    files = list_drive_files(g_token, CUDYR_2026_FOLDER)
    print(f"  → {len(files)} archivos encontrados")
    
    cudyr = {}
    for f in files:
        mes_nombre, mes_num = extract_month_from_name(f["name"])
        if not mes_nombre:
            continue
        
        print(f"  Procesando CUDYR {mes_nombre}...")
        datos = get_cudyr_via_sheets_api(g_token, f["id"])
        if datos:
            cudyr[mes_nombre] = datos
            print(f"  → Críticos: {datos['criticos_pct']}% | Medios: {datos['medios_pct']}% | Básicos: {datos['basicos_pct']}%")
        else:
            print(f"  ⚠ No se pudieron extraer datos de {f['name']}")
    
    return cudyr

# ══════════════════════════════════════════════════════════════════════
# CENSO — parsear meses del Excel SharePoint
# ══════════════════════════════════════════════════════════════════════
def parse_mes(ws):
    datos_diarios = []
    dotacion = 29

    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and "DOTACION" in str(cell).upper():
                idx = list(row).index(cell)
                if idx + 1 < len(row) and row[idx+1]:
                    try: dotacion = int(row[idx+1])
                    except: pass

        first = row[0]
        if first is None:
            continue
        fecha = None
        if isinstance(first, datetime):
            fecha = first
        else:
            try: fecha = datetime.strptime(str(first).strip(), "%d/%m/%Y")
            except: pass
        if fecha is None:
            continue

        def si(v):
            try: return int(v) if v is not None else 0
            except: return 0

        fila = list(row)
        datos_diarios.append({
            "fecha": fecha.strftime("%Y-%m-%d"),
            "existencia": si(fila[1]),
            "ing_urgencia": si(fila[2]),
            "ing_aps": si(fila[3]),
            "ing_cae": si(fila[4]),
            "ing_otros_hosp": si(fila[5]),
            "ing_otra_proc": si(fila[6]),
            "ing_mismo_serv": si(fila[7]),
            "total_ingresos": si(fila[8]),
            "egr_alta": si(fila[9]),
            "egr_traslado": si(fila[10]),
            "egr_fallecidos": si(fila[11]),
            "total_egresos": si(fila[12]),
            "mismo_dia": si(fila[13]),
            "camas_disponibles": si(fila[14]),
            "camas_ocupadas": si(fila[15]),
            "dias_estada": si(fila[16]),
        })

    if not datos_diarios:
        return None

    total_ingresos  = sum(d["total_ingresos"] for d in datos_diarios)
    total_egresos   = sum(d["total_egresos"] for d in datos_diarios)
    total_fallecidos= sum(d["egr_fallecidos"] for d in datos_diarios)
    total_dias_estada= sum(d["dias_estada"] for d in datos_diarios)
    dias = len(datos_diarios)
    ocup_prom = round(sum(d["camas_ocupadas"] for d in datos_diarios)/dias, 1) if dias else 0
    pct_ocup  = round(ocup_prom/dotacion*100, 1) if dotacion else 0
    estada    = round(total_dias_estada/total_egresos, 1) if total_egresos else 0

    return {
        "dias": datos_diarios,
        "resumen": {
            "dotacion": dotacion,
            "total_ingresos": total_ingresos,
            "total_egresos": total_egresos,
            "total_fallecidos": total_fallecidos,
            "ocupacion_promedio": ocup_prom,
            "porcentaje_ocupacion": pct_ocup,
            "estada_promedio": estada,
            "ing_urgencia": sum(d["ing_urgencia"] for d in datos_diarios),
            "ing_aps": sum(d["ing_aps"] for d in datos_diarios),
            "ing_otros_hosp": sum(d["ing_otros_hosp"] for d in datos_diarios),
        }
    }

def parse_sociosanitarios(ws):
    casos = []
    hoy = date.today()
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip().upper() == "OTROS":
            header_found = True
            continue
        if not header_found:
            continue
        if str(row[0]).strip().upper() != "FIJO":
            continue

        def sv(v): return str(v).strip() if v is not None else ""

        nombre_completo = f"{sv(row[3])} {sv(row[1])} {sv(row[2])}".strip()
        fecha_ing_raw = row[14]
        fecha_ing_str = ""
        dias_hosp = None

        if fecha_ing_raw:
            if isinstance(fecha_ing_raw, datetime):
                fi = fecha_ing_raw.date()
            else:
                try: fi = datetime.strptime(str(fecha_ing_raw).strip(), "%d/%m/%Y").date()
                except: fi = None
            if fi:
                fecha_ing_str = fi.strftime("%d/%m/%Y")
                dias_hosp = (hoy - fi).days

        if nombre_completo:
            casos.append({
                "nombre": nombre_completo,
                "rut": sv(row[5]),
                "edad": sv(row[6]),
                "fecha_ingreso": fecha_ing_str,
                "dias_hospitalizados": dias_hosp,
                "sala": sv(row[12]),
                "cama": sv(row[13]),
                "procedencia": sv(row[8]),
                "prevision": sv(row[9]),
            })

    casos.sort(key=lambda x: x["dias_hospitalizados"] or 0, reverse=True)
    return casos

# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
def main():
    print("─── Microsoft / SharePoint ───")
    print("Obteniendo token MS...")
    ms_token = get_ms_token()
    print("Descargando Excel censo...")
    content = get_sharepoint_file(ms_token)
    print("Procesando Excel...")
    wb = load_workbook(BytesIO(content), data_only=True)

    resultado = {
        "actualizado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "meses": {},
        "sociosanitarios": [],
        "cudyr": {}
    }

    for sheet_name in wb.sheetnames:
        nombre_upper = sheet_name.strip().upper()
        if nombre_upper in MESES_VALIDOS:
            print(f"  Procesando mes: {sheet_name}")
            datos = parse_mes(wb[sheet_name])
            if datos:
                resultado["meses"][nombre_upper] = datos
        elif nombre_upper == "DATOS":
            print("  Procesando casos sociosanitarios...")
            casos = parse_sociosanitarios(wb[sheet_name])
            resultado["sociosanitarios"] = casos
            print(f"  → {len(casos)} casos FIJO")

    print("\n─── Google Drive / CUDYR ───")
    try:
        g_token = get_google_token()
        cudyr = fetch_cudyr_data(g_token)
        resultado["cudyr"] = cudyr
        print(f"✓ CUDYR: {list(cudyr.keys())}")
    except Exception as e:
        print(f"⚠ Error CUDYR: {e}")
        resultado["cudyr"] = {}

    os.makedirs("data", exist_ok=True)
    with open("data/censo.json", "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Listo. Meses censo: {list(resultado['meses'].keys())}")
    print(f"✓ Casos sociosanitarios: {len(resultado['sociosanitarios'])}")
    print(f"✓ CUDYR meses: {list(resultado['cudyr'].keys())}")

if __name__ == "__main__":
    main()
